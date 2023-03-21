from datetime import datetime, timedelta

import openpyxl

from db import DataBaseManager


class XlsxParser:
    def __init__(self, xlsx_file):
        self.workbook = openpyxl.load_workbook(xlsx_file)
        self.worksheet = self.workbook.active
        self.start_row_for_parse = 3
        self.end_column_for_parse = self.worksheet.max_column - 2
        self.db_manager = DataBaseManager()

    def get_data_from_file(self):
        for r in range(self.start_row_for_parse, self.worksheet.max_row):
            row = list()
            for c in self.worksheet.iter_cols(1, self.end_column_for_parse):
                row.append(c[r].value)
            row.append(datetime.today() - timedelta(days=r))
            self.create_company_forecast(row)

    def create_company_forecast(self, row):
        self.db_manager.add_company_forecast(row)

    def get_total(self):
        return self.db_manager.get_total()


if __name__ == '__main__':
    parser = XlsxParser("./Приложение.xlsx")
    parser.get_data_from_file()
    print(parser.get_total())
